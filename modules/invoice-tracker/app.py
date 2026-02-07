"""
Invoice Payment Tracker - Main Application
Handles invoice upload, OCR processing, and Google Sheets integration
"""

import os
import io
import json
import logging
import zipfile
import gc
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from functools import wraps
from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, flash, session, make_response, send_file
)
from flask_cors import CORS
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

# Import our custom modules
from invoice_processor import InvoiceProcessor, process_invoice
from sheets_manager import (
    SheetsManager, SheetsManagerError, AuthenticationError,
    save_to_sheets, save_payment_details as save_payment_to_sheets,
    get_invoices, get_payments
)

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
CORS(app, origins=['https://coreworker-landing.onrender.com', 'http://localhost:*'])  # Restrict CORS

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', os.urandom(24))
app.config['UPLOAD_FOLDER'] = os.path.abspath(os.getenv('UPLOAD_FOLDER', 'uploads'))
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', 16777216))  # 16MB
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'png', 'jpg', 'jpeg'}
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('FLASK_ENV', 'development') == 'production'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


# ========== Security Headers ==========

@app.after_request
def add_security_headers(response):
    """Add security headers to all responses"""
    # Prevent clickjacking
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    # Prevent MIME type sniffing
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # XSS Protection (legacy browsers)
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # Referrer Policy
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    # Content Security Policy
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com https://cdn.jsdelivr.net; "
        "img-src 'self' data: https:; "
        "connect-src 'self' https://sheets.googleapis.com https://www.googleapis.com https://drive.googleapis.com; "
        "frame-ancestors 'self';"
    )
    return response


# ========== Demo Mode Configuration ==========

# Set DEMO_MODE=true in environment to disable password protection
DEMO_MODE = os.environ.get('DEMO_MODE', 'true').lower() == 'true'


# ========== User Profiles ==========

# Profile definitions with role-based module access
USER_PROFILES = {
    1: {
        'id': 1,
        'name': 'Rob Daykin',
        'email': 'robert.daykin84@gmail.com',
        'role': 'Admin',
        'avatar': 'RD',
        'color': '#54A6ED',
        'modules': ['invoice_tracker', 'client_onboarding', 'capital_calls', 'distributions']
    },
    2: {
        'id': 2,
        'name': 'Sarah Mitchell',
        'email': 'sarah.mitchell@fundadmin.com',
        'role': 'Fund Administrator',
        'avatar': 'SM',
        'color': '#10B981',
        'modules': ['invoice_tracker', 'client_onboarding']
    },
    3: {
        'id': 3,
        'name': 'James Chen',
        'email': 'james.chen@fundadmin.com',
        'role': 'Compliance Officer',
        'avatar': 'JC',
        'color': '#8B5CF6',
        'modules': ['client_onboarding']
    },
    4: {
        'id': 4,
        'name': 'Emma Rodriguez',
        'email': 'emma.rodriguez@fundadmin.com',
        'role': 'Operations Manager',
        'avatar': 'ER',
        'color': '#F59E0B',
        'modules': ['invoice_tracker', 'capital_calls', 'distributions']
    }
}


@app.context_processor
def inject_demo_mode():
    """Inject demo_mode flag, user profile, and configuration status into all templates"""
    config_status = {
        'google_sheets_configured': bool(os.environ.get('GOOGLE_SHEET_ID')),
        'anthropic_configured': bool(os.environ.get('ANTHROPIC_API_KEY')),
        'google_credentials_configured': bool(os.environ.get('GOOGLE_APPLICATION_CREDENTIALS') or os.environ.get('GOOGLE_SERVICE_ACCOUNT_JSON'))
    }
    # Get current user profile from session
    current_user = None
    if session.get('user_id'):
        current_user = USER_PROFILES.get(session['user_id'])
    return {
        'demo_mode': DEMO_MODE,
        'config_status': config_status,
        'current_user': current_user
    }


# ========== Custom Jinja2 Filters ==========

def sanitize_error_value(value, default=''):
    """
    Filter to sanitize AI extraction error values like #ERROR!, N/A, etc.
    Returns the default value if the input contains error indicators.
    """
    if value is None:
        return default

    value_str = str(value).strip()

    # List of error indicators that should be replaced
    error_indicators = [
        '#ERROR!', '#error!', '#ERROR', '#error',
        '#N/A', '#n/a', '#NA', '#na',
        '#VALUE!', '#value!',
        '#REF!', '#ref!',
        '#DIV/0!', '#div/0!',
        '#NULL!', '#null!',
        '#NAME?', '#name?',
        'N/A', 'n/a', 'NA', 'na',
        'undefined', 'null', 'None',
        'ERROR', 'Error', 'error'
    ]

    # Check if value matches any error indicator
    if value_str in error_indicators or value_str.startswith('#'):
        return default

    return value_str

# Register the filter with Jinja2
app.jinja_env.filters['sanitize'] = sanitize_error_value


# ========== Authentication ==========

def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('authenticated'):
            if DEMO_MODE:
                # In demo mode, redirect to profile selection instead of auto-auth
                return redirect(url_for('login', next=request.url))
            else:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function


# ========== Helper Functions ==========

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def get_file_extension(filename):
    """Get file extension"""
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''


def generate_unique_filename(original_filename):
    """Generate a unique filename with timestamp"""
    ext = get_file_extension(original_filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = secure_filename(original_filename.rsplit('.', 1)[0])
    return f"{base_name}_{timestamp}.{ext}"


def handle_errors(f):
    """Decorator for handling errors in routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except AuthenticationError as e:
            logger.error(f"Authentication error: {e}")
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'success': False,
                    'error': 'Google Sheets authentication failed. Please check your credentials.',
                    'error_type': 'authentication'
                }), 401
            flash('Google Sheets authentication failed. Please check your credentials.', 'danger')
            return redirect(url_for('index'))
        except SheetsManagerError as e:
            logger.error(f"Sheets manager error: {e}")
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'success': False,
                    'error': str(e),
                    'error_type': 'sheets_error'
                }), 500
            flash(f'Google Sheets error: {str(e)}', 'danger')
            return redirect(url_for('index'))
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'success': False,
                    'error': 'An unexpected error occurred',
                    'error_type': 'server_error'
                }), 500
            flash('An unexpected error occurred. Please try again.', 'danger')
            return redirect(url_for('index'))
    return decorated_function


def get_sheets_manager():
    """Get or create a SheetsManager instance"""
    try:
        return SheetsManager()
    except Exception as e:
        logger.error(f"Failed to initialize SheetsManager: {e}")
        raise


# ========== Authentication Routes ==========

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page with profile selection"""
    # Already logged in - redirect to hub
    if session.get('authenticated'):
        return redirect(url_for('hub'))

    if request.method == 'POST':
        if DEMO_MODE:
            # Profile selection login
            profile_id = request.form.get('profile_id')
            if profile_id:
                profile_id = int(profile_id)
                profile = USER_PROFILES.get(profile_id)
                if profile:
                    session['authenticated'] = True
                    session['user_id'] = profile['id']
                    session['user_name'] = profile['name']
                    session['user_role'] = profile['role']
                    session['user_avatar'] = profile['avatar']
                    session['user_color'] = profile['color']
                    session['user_modules'] = profile['modules']
                    session.permanent = True
                    logger.info(f"User logged in: {profile['name']} ({profile['role']})")
                    next_url = request.args.get('next')
                    return redirect(next_url or url_for('hub'))
            flash('Please select a profile to continue.', 'warning')
        else:
            # Password-based login (non-demo mode)
            password = request.form.get('password', '')
            app_password = os.environ.get('APP_PASSWORD', '')

            if not app_password:
                flash('APP_PASSWORD not configured. Please set it in your .env file.', 'error')
                return render_template('login.html', profiles=USER_PROFILES, demo_mode=DEMO_MODE)

            if password == app_password:
                # In non-demo mode with password, default to Admin profile
                admin_profile = USER_PROFILES[1]
                session['authenticated'] = True
                session['user_id'] = admin_profile['id']
                session['user_name'] = admin_profile['name']
                session['user_role'] = admin_profile['role']
                session['user_avatar'] = admin_profile['avatar']
                session['user_color'] = admin_profile['color']
                session['user_modules'] = admin_profile['modules']
                session.permanent = True
                logger.info("User logged in via password (Admin)")
                next_url = request.args.get('next')
                return redirect(next_url or url_for('hub'))
            else:
                flash('Invalid password. Please try again.', 'error')
                logger.warning("Failed login attempt")

    return render_template('login.html', profiles=USER_PROFILES, demo_mode=DEMO_MODE)


@app.route('/logout')
def logout():
    """Logout and clear session"""
    session.clear()
    return redirect(url_for('login'))


# ========== Main Routes ==========

@app.route('/')
@handle_errors
@login_required
def index():
    """Root redirect to hub"""
    return redirect(url_for('hub'))


@app.route('/hub')
@handle_errors
@login_required
def hub():
    """Hub landing page with module navigation"""
    return render_template('hub.html')


@app.route('/invoice-dashboard')
@handle_errors
@login_required
def invoice_dashboard():
    """Invoice dashboard - original home page"""
    try:
        manager = get_sheets_manager()
        stats = manager.get_invoice_stats()
        recent_invoices = manager.get_recent_invoices(limit=5)
        payment_details = manager.get_all_payment_details()
    except Exception as e:
        logger.warning(f"Could not load dashboard data: {e}")
        stats = {
            'total_invoices': 0,
            'paid': 0,
            'pending': 0,
            'overdue': 0
        }
        recent_invoices = []
        payment_details = []

    return render_template('index.html', stats=stats, recent_invoices=recent_invoices, payment_details=payment_details)


@app.route('/dashboard')
@handle_errors
@login_required
def dashboard():
    """Dashboard page showing all invoices"""
    try:
        manager = get_sheets_manager()
        invoices = manager.get_all_invoices()
        stats = manager.get_invoice_stats()
        payment_details = manager.get_all_payment_details()
    except Exception as e:
        logger.warning(f"Could not load dashboard data: {e}")
        invoices = []
        stats = {}
        payment_details = []

    return render_template('index.html', stats=stats, recent_invoices=invoices, payment_details=payment_details)


# ========== Upload Routes ==========

@app.route('/upload', methods=['GET', 'POST'])
@handle_errors
@login_required
def upload():
    """Handle invoice upload and processing"""
    if request.method == 'POST':
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.best == 'application/json'

        # Check if file was uploaded (support both 'file' and 'invoice' field names)
        file = request.files.get('file') or request.files.get('invoice')

        if not file:
            logger.warning("Upload attempted with no file")
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': 'No file uploaded'
                }), 400
            flash('No file uploaded', 'danger')
            return redirect(url_for('upload'))

        if file.filename == '':
            logger.warning("Upload attempted with empty filename")
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': 'No file selected'
                }), 400
            flash('No file selected', 'danger')
            return redirect(url_for('upload'))

        if not allowed_file(file.filename):
            logger.warning(f"Upload attempted with invalid file type: {file.filename}")
            error_msg = f'Invalid file type. Allowed types: {", ".join(app.config["ALLOWED_EXTENSIONS"])}'
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 400
            flash(error_msg, 'danger')
            return redirect(url_for('upload'))

        # Generate unique filename and save
        filename = generate_unique_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        try:
            file.save(filepath)
            logger.info(f"File saved: {filepath}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': 'Failed to save uploaded file'
                }), 500
            flash('Failed to save uploaded file', 'danger')
            return redirect(url_for('upload'))

        # Process invoice with Claude API
        try:
            logger.info(f"Processing invoice: {filepath}")
            extracted_data = process_invoice(filepath)

            # Check for processing errors
            if extracted_data.get('error'):
                logger.warning(f"Invoice processing returned error: {extracted_data.get('error')}")
                flash(f"Warning: {extracted_data.get('error')}", 'warning')

            logger.info(f"Invoice processed successfully: {extracted_data.get('invoice_number', 'N/A')}")

            # Check for duplicate invoice (same supplier + invoice number)
            try:
                manager = get_sheets_manager()
                existing_invoices = manager.get_all_invoices()
                new_invoice_num = extracted_data.get('invoice_number', '').strip().lower()
                new_supplier = extracted_data.get('supplier_name', '').strip().lower()

                for existing in existing_invoices:
                    existing_num = str(existing.get('invoice_number', '')).strip().lower()
                    existing_supplier = str(existing.get('supplier_name', '')).strip().lower()

                    if new_invoice_num and new_supplier and new_invoice_num == existing_num and new_supplier == existing_supplier:
                        extracted_data['_potential_duplicate'] = True
                        extracted_data['_duplicate_info'] = {
                            'invoice_number': existing.get('invoice_number'),
                            'supplier_name': existing.get('supplier_name'),
                            'invoice_date': existing.get('invoice_date'),
                            'amount': existing.get('amount')
                        }
                        logger.warning(f"Potential duplicate invoice detected: {new_invoice_num} from {new_supplier}")
                        break
            except Exception as e:
                logger.warning(f"Could not check for duplicates: {e}")

            # Upload file to Google Drive for persistent storage
            try:
                manager = get_sheets_manager()
                drive_result = manager.upload_file_to_drive(filepath, filename)
                if drive_result.get('success'):
                    extracted_data['file_id'] = drive_result.get('file_id')
                    logger.info(f"File uploaded to Google Drive: {drive_result.get('file_id')}")
                else:
                    logger.warning(f"Failed to upload to Google Drive: {drive_result.get('message')}")
            except Exception as e:
                logger.warning(f"Could not upload to Google Drive: {e}")

            # Save extracted data to JSON file (session cookie is too small)
            json_filename = filename.rsplit('.', 1)[0] + '_data.json'
            json_filepath = os.path.join(app.config['UPLOAD_FOLDER'], json_filename)
            with open(json_filepath, 'w') as f:
                json.dump(extracted_data, f)
            logger.info(f"Extracted data saved to: {json_filepath}")

            # Store only the filename in session (small enough for cookie)
            session['invoice_filename'] = filename

            # Free up memory after processing
            gc.collect()

            # Always redirect to review page for verification
            if is_ajax:
                return jsonify({
                    'success': True,
                    'data': extracted_data,
                    'filename': filename,
                    'redirect': url_for('review', filename=filename)
                })

            # Direct redirect for form submission
            return redirect(url_for('review', filename=filename))

        except Exception as e:
            logger.error(f"Invoice processing failed: {e}", exc_info=True)
            error_msg = f'Failed to process invoice: {str(e)}'
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 500
            flash(error_msg, 'danger')
            return redirect(url_for('upload'))

    # GET request - show upload form
    return render_template('upload.html')


# ========== Review Routes ==========

def convert_date_for_form(date_str):
    """Convert date from DD/MM/YYYY to YYYY-MM-DD format for HTML date inputs"""
    if not date_str:
        return ''
    # If already in YYYY-MM-DD format, return as is
    if len(date_str) == 10 and date_str[4] == '-' and date_str[7] == '-':
        return date_str
    # Try to parse DD/MM/YYYY format
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            day, month, year = parts
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    except (ValueError, IndexError):
        pass
    return date_str


@app.route('/review')
@app.route('/review/<filename>')
@handle_errors
@login_required
def review(filename=None):
    """Review and edit extracted invoice data"""

    # Check if editing an existing invoice from Google Sheets
    invoice_number = request.args.get('invoice_number')
    if invoice_number:
        logger.info(f"Loading existing invoice for editing: {invoice_number}")
        try:
            manager = get_sheets_manager()
            invoice = manager.get_invoice_by_number(invoice_number)

            if not invoice:
                flash(f'Invoice {invoice_number} not found.', 'warning')
                return redirect(url_for('dashboard'))

            # Get associated payment details
            payment_details = None
            all_payments = manager.get_all_payment_details()
            for payment in all_payments:
                if payment.get('invoice_number') == invoice_number:
                    payment_details = payment
                    break

            # Format invoice data for the template (convert dates for form inputs)
            invoice_data = {
                'invoice_number': invoice.get('invoice_number', ''),
                'supplier_name': invoice.get('supplier_name', ''),
                'contact_email': invoice.get('contact_email', ''),
                'contact_phone': invoice.get('contact_phone', ''),
                'invoice_date': convert_date_for_form(invoice.get('invoice_date', '')),
                'due_date': convert_date_for_form(invoice.get('due_date', '')),
                'amount': invoice.get('amount', 0),
                'currency': invoice.get('currency', 'GBP'),
                'status': invoice.get('status', ''),
                'payment_date': convert_date_for_form(invoice.get('payment_date', '')),
                'notes': invoice.get('notes', ''),
                'payment_details': payment_details,
                '_row_id': invoice.get('id')  # Row number for updates
            }

            return render_template('review.html', invoice=invoice_data, is_editing=True)

        except Exception as e:
            logger.error(f"Error loading invoice for editing: {e}")
            flash(f'Error loading invoice: {str(e)}', 'danger')
            return redirect(url_for('dashboard'))

    # Try to get data from JSON file or query params (for new uploads)
    filename = filename or session.get('invoice_filename') or request.args.get('filename')
    invoice_data = {}

    # Handle queue of multiple invoices
    queue_param = request.args.get('queue', '')
    queue = queue_param.split(',') if queue_param else []
    current_index = 0
    total_invoices = len(queue) if queue else 1

    if filename and queue:
        try:
            current_index = queue.index(filename)
        except ValueError:
            current_index = 0

    if filename:
        # Try to load extracted data from JSON file
        json_filename = filename.rsplit('.', 1)[0] + '_data.json'
        json_filepath = os.path.join(app.config['UPLOAD_FOLDER'], json_filename)

        if os.path.exists(json_filepath):
            try:
                with open(json_filepath, 'r') as f:
                    invoice_data = json.load(f)
                logger.info(f"Loaded extracted data from: {json_filepath}")
            except Exception as e:
                logger.warning(f"Could not load JSON data: {e}")

        invoice_data['file_path'] = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not invoice_data or not invoice_data.get('invoice_number'):
        flash('No invoice data found. Please upload an invoice first.', 'warning')
        return redirect(url_for('upload'))

    return render_template('review.html',
                           invoice=invoice_data,
                           is_editing=False,
                           queue=queue,
                           current_index=current_index,
                           total_invoices=total_invoices,
                           current_filename=filename)


@app.route('/save-invoice', methods=['POST'])
@handle_errors
@login_required
def save_invoice():
    """Save invoice data to Google Sheets (create new or update existing)"""
    # Get form data
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form.to_dict()

    # Check if this is an edit of an existing invoice
    is_editing = data.get('is_editing') == 'true' or data.get('is_editing') == True
    row_id = data.get('_row_id')

    # Validate required fields
    required_fields = ['invoice_number', 'supplier_name', 'amount']
    missing_fields = [f for f in required_fields if not data.get(f)]

    if missing_fields:
        return jsonify({
            'success': False,
            'error': f'Missing required fields: {", ".join(missing_fields)}'
        }), 400

    # Clean and prepare data
    invoice_data = {
        'invoice_number': data.get('invoice_number', '').strip(),
        'supplier_name': data.get('supplier_name', '').strip(),
        'contact_email': data.get('contact_email', '').strip(),
        'contact_phone': data.get('contact_phone', '').strip(),
        'invoice_date': data.get('invoice_date', ''),
        'due_date': data.get('due_date', ''),
        'amount': float(data.get('amount', 0) or 0),
        'currency': data.get('currency', 'GBP').upper(),
        'status': data.get('status', 'Pending Review'),
        'payment_date': data.get('payment_date', ''),
        'notes': data.get('notes', '').strip(),
        'file_id': data.get('file_id', '')
    }

    # Extract payment details if present
    payment_details = None
    if data.get('payment_details') and isinstance(data.get('payment_details'), dict):
        payment_details = data.get('payment_details')
    else:
        # Check for individual payment fields in the form data
        payment_fields = ['beneficiary_account_name', 'account_number', 'iban',
                         'sort_code', 'swift_code', 'bank_name', 'bank_address', 'payment_reference']
        has_payment_data = any(data.get(f) for f in payment_fields)
        if has_payment_data:
            payment_details = {
                'beneficiary_account_name': data.get('beneficiary_account_name', '').strip(),
                'account_number': data.get('account_number', '').strip(),
                'iban': data.get('iban', '').strip().upper(),
                'sort_code': data.get('sort_code', '').strip(),
                'swift_code': data.get('swift_code', '').strip().upper(),
                'bank_name': data.get('bank_name', '').strip(),
                'bank_address': data.get('bank_address', '').strip(),
                'payment_reference': data.get('payment_reference', '').strip()
            }

    # Save to Google Sheets (update or create)
    try:
        manager = get_sheets_manager()

        if is_editing and row_id:
            # Update existing invoice
            row_number = int(row_id)
            result = manager.update_invoice(row_number, invoice_data)
            logger.info(f"Updated existing invoice: {invoice_data['invoice_number']} at row {row_number}")
        else:
            # Create new invoice
            result = save_to_sheets(invoice_data)

        if result.get('success'):
            # Handle payment details
            has_payment_info = payment_details and any(payment_details.values())
            payment_data = {
                'invoice_number': invoice_data['invoice_number'],
                'supplier_name': invoice_data['supplier_name'],
                'beneficiary_account_name': payment_details.get('beneficiary_account_name', '') if payment_details else '',
                'account_number': payment_details.get('account_number', '') if payment_details else '',
                'iban': payment_details.get('iban', '') if payment_details else '',
                'sort_code': payment_details.get('sort_code', '') if payment_details else '',
                'swift_code': payment_details.get('swift_code', '') if payment_details else '',
                'bank_name': payment_details.get('bank_name', '') if payment_details else '',
                'bank_address': payment_details.get('bank_address', '') if payment_details else '',
                'payment_reference': payment_details.get('payment_reference', '') if payment_details else '',
                'status': 'Auto-populated' if has_payment_info else 'Pending Details',
                'upload_date': '',
                'notes': f'Auto-populated from invoice {invoice_data["invoice_number"]}' if has_payment_info else f'Awaiting payment details - from invoice {invoice_data["invoice_number"]}'
            }

            try:
                if is_editing:
                    # Check if payment details exist for this invoice and update them
                    existing_payment = None
                    all_payments = manager.get_all_payment_details()
                    for p in all_payments:
                        if p.get('invoice_number') == invoice_data['invoice_number']:
                            existing_payment = p
                            break

                    if existing_payment:
                        manager.update_payment_details(existing_payment['id'], payment_data)
                        logger.info(f"Payment details updated for invoice: {invoice_data['invoice_number']}")
                    else:
                        save_payment_to_sheets(payment_data)
                        logger.info(f"Payment details created for invoice: {invoice_data['invoice_number']}")
                else:
                    save_payment_to_sheets(payment_data)
                    logger.info(f"Payment details saved for supplier: {invoice_data['supplier_name']}")
            except Exception as e:
                logger.warning(f"Failed to save payment details: {e}")

            # Clear session data (only for new uploads)
            if not is_editing:
                session.pop('extracted_invoice', None)
                session.pop('invoice_filename', None)

            success_msg = 'Invoice updated successfully!' if is_editing else 'Invoice saved successfully!'
            logger.info(f"Invoice {'updated' if is_editing else 'saved'} to sheets: {invoice_data['invoice_number']}")

            # Handle queue navigation - go to next invoice if there's more in queue
            queue_param = data.get('queue', '')
            current_index = int(data.get('current_index', 0))
            queue = queue_param.split(',') if queue_param else []

            if queue and current_index < len(queue) - 1:
                # There are more invoices to review
                next_index = current_index + 1
                next_filename = queue[next_index]
                next_url = url_for('review', filename=next_filename) + f'?queue={queue_param}'
                remaining = len(queue) - next_index
                flash(f'{success_msg} {remaining} more invoice(s) to review.', 'success')

                if request.is_json:
                    return jsonify({
                        'success': True,
                        'message': success_msg,
                        'redirect': next_url,
                        'remaining': remaining
                    })
                return redirect(next_url)

            # No more invoices in queue, go to dashboard
            if request.is_json:
                return jsonify({
                    'success': True,
                    'message': success_msg,
                    'redirect': url_for('dashboard')
                })
            flash(success_msg, 'success')
            return redirect(url_for('dashboard'))
        else:
            error_msg = result.get('message', 'Failed to save invoice')
            logger.error(f"Failed to save invoice: {error_msg}")

            if request.is_json:
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 500
            flash(error_msg, 'danger')
            return redirect(url_for('review'))

    except Exception as e:
        logger.error(f"Error saving invoice: {e}", exc_info=True)
        if request.is_json:
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500
        flash(f'Error saving invoice: {str(e)}', 'danger')
        return redirect(url_for('review'))


# ========== Payment Details Routes ==========

@app.route('/payment-details', methods=['GET', 'POST'])
@handle_errors
@login_required
def payment_details():
    """Manage payment details"""
    if request.method == 'POST':
        # Handle form submission
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()

        # Use new_supplier if provided, otherwise use selected supplier
        supplier_name = data.get('new_supplier', '').strip() or data.get('supplier_name', '').strip()

        if not supplier_name:
            if request.is_json:
                return jsonify({
                    'success': False,
                    'error': 'Supplier name is required'
                }), 400
            flash('Supplier name is required', 'danger')
            return redirect(url_for('payment_details'))

        # Prepare payment data
        payment_data = {
            'invoice_number': data.get('invoice_number', '').strip(),
            'supplier_name': supplier_name,
            'beneficiary_account_name': data.get('beneficiary_account_name', '').strip(),
            'account_number': data.get('account_number', '').strip(),
            'iban': data.get('iban', '').strip().upper(),
            'sort_code': data.get('sort_code', '').strip(),
            'swift_code': data.get('swift_code', '').strip().upper(),
            'bank_name': data.get('bank_name', '').strip(),
            'bank_address': data.get('bank_address', '').strip(),
            'payment_reference': data.get('payment_reference', '').strip(),
            'status': data.get('status', 'pending'),
            'upload_date': data.get('upload_date', ''),
            'notes': data.get('notes', '').strip()
        }

        # Save to Google Sheets
        try:
            result = save_payment_to_sheets(payment_data)

            if result.get('success'):
                logger.info(f"Payment details saved for: {supplier_name}")

                if request.is_json:
                    return jsonify({
                        'success': True,
                        'message': 'Payment details saved successfully'
                    })
                flash('Payment details saved successfully!', 'success')
                return redirect(url_for('payment_details'))
            else:
                error_msg = result.get('message', 'Failed to save payment details')
                logger.error(f"Failed to save payment details: {error_msg}")

                if request.is_json:
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    }), 500
                flash(error_msg, 'danger')
                return redirect(url_for('payment_details'))

        except Exception as e:
            logger.error(f"Error saving payment details: {e}", exc_info=True)
            if request.is_json:
                return jsonify({
                    'success': False,
                    'error': str(e)
                }), 500
            flash(f'Error saving payment details: {str(e)}', 'danger')
            return redirect(url_for('payment_details'))

    # GET request - show payment details page
    try:
        manager = get_sheets_manager()
        payment_list = manager.get_all_payment_details()
        suppliers = manager.get_unique_suppliers()
        invoices = manager.get_all_invoices()
    except Exception as e:
        logger.warning(f"Could not load payment details: {e}")
        payment_list = []
        suppliers = []
        invoices = []

    return render_template(
        'payment_details.html',
        payment_details=payment_list,
        suppliers=suppliers,
        invoices=invoices
    )


@app.route('/save-payment-details', methods=['POST'])
@handle_errors
@login_required
def save_payment_details_route():
    """Alternative endpoint for saving payment details"""
    return payment_details()


@app.route('/download-payment-report')
@handle_errors
@login_required
def download_payment_report():
    """Download all payment details as an Excel file"""
    try:
        manager = get_sheets_manager()
        payments = manager.get_all_payment_details()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Details"

        # Define headers
        headers = [
            'Invoice Number',
            'Supplier Name',
            'Beneficiary Account Name',
            'Account Number',
            'IBAN',
            'Sort Code',
            'SWIFT/BIC Code',
            'Bank Name',
            'Bank Address',
            'Payment Reference',
            'Status',
            'Upload Date',
            'Notes'
        ]

        # Write headers with bold formatting
        bold_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold_font

        # Write data rows
        for row_num, payment in enumerate(payments, 2):
            ws.cell(row=row_num, column=1, value=payment.get('invoice_number', ''))
            ws.cell(row=row_num, column=2, value=payment.get('supplier_name', ''))
            ws.cell(row=row_num, column=3, value=payment.get('beneficiary_account_name', ''))
            ws.cell(row=row_num, column=4, value=payment.get('account_number', ''))
            ws.cell(row=row_num, column=5, value=payment.get('iban', ''))
            ws.cell(row=row_num, column=6, value=payment.get('sort_code', ''))
            ws.cell(row=row_num, column=7, value=payment.get('swift_code', ''))
            ws.cell(row=row_num, column=8, value=payment.get('bank_name', ''))
            ws.cell(row=row_num, column=9, value=payment.get('bank_address', ''))
            ws.cell(row=row_num, column=10, value=payment.get('payment_reference', ''))
            ws.cell(row=row_num, column=11, value=payment.get('status', ''))
            ws.cell(row=row_num, column=12, value=payment.get('upload_date', ''))
            ws.cell(row=row_num, column=13, value=payment.get('notes', ''))

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(payments) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'Payment_Report_{today}.xlsx'

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        logger.info(f"Payment report downloaded: {len(payments)} records")
        return response

    except Exception as e:
        logger.error(f"Error generating payment report: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to generate report: {str(e)}'
        }), 500


@app.route('/api/invoices/download-report', methods=['POST'])
@handle_errors
@login_required
def download_invoice_report():
    """Download selected invoices as an Excel file with payment details and totals"""
    try:
        # Get selected invoice numbers from request
        data = request.get_json() or {}
        selected_invoices = data.get('invoice_numbers', [])

        manager = get_sheets_manager()
        all_invoices = manager.get_all_invoices()
        all_payments = manager.get_all_payment_details()

        # Filter invoices if specific ones were selected
        if selected_invoices:
            invoices = [inv for inv in all_invoices if inv.get('invoice_number') in selected_invoices]
        else:
            invoices = all_invoices

        # Create payment lookup by invoice number
        payment_lookup = {}
        for payment in all_payments:
            inv_num = payment.get('invoice_number', '')
            if inv_num:
                payment_lookup[inv_num] = payment

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Report"

        # Define headers
        headers = [
            'Invoice Number',
            'Supplier Name',
            'Contact Email',
            'Contact Phone',
            'Invoice Date',
            'Due Date',
            'Amount',
            'Currency',
            'Status',
            'Payment Date',
            'Notes',
            # Payment Details
            'Beneficiary Account Name',
            'Bank Name',
            'Account Number',
            'Sort Code',
            'IBAN',
            'SWIFT/BIC Code',
            'Payment Reference',
            'Bank Address'
        ]

        # Write headers with bold formatting
        bold_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold_font

        # Track total amount
        total_amount = 0.0

        # Write data rows
        for row_num, invoice in enumerate(invoices, 2):
            inv_num = invoice.get('invoice_number', '')
            payment = payment_lookup.get(inv_num, {})
            amount = invoice.get('amount', 0) or 0

            # Try to convert amount to float
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                amount = 0.0

            total_amount += amount

            # Invoice details
            ws.cell(row=row_num, column=1, value=inv_num)
            ws.cell(row=row_num, column=2, value=invoice.get('supplier_name', ''))
            ws.cell(row=row_num, column=3, value=invoice.get('contact_email', ''))
            ws.cell(row=row_num, column=4, value=invoice.get('contact_phone', ''))
            ws.cell(row=row_num, column=5, value=invoice.get('invoice_date', ''))
            ws.cell(row=row_num, column=6, value=invoice.get('due_date', ''))
            ws.cell(row=row_num, column=7, value=amount)
            ws.cell(row=row_num, column=8, value=invoice.get('currency', 'GBP'))
            ws.cell(row=row_num, column=9, value=invoice.get('status', ''))
            ws.cell(row=row_num, column=10, value=invoice.get('payment_date', ''))
            ws.cell(row=row_num, column=11, value=invoice.get('notes', ''))

            # Payment details
            ws.cell(row=row_num, column=12, value=payment.get('beneficiary_account_name', ''))
            ws.cell(row=row_num, column=13, value=payment.get('bank_name', ''))
            ws.cell(row=row_num, column=14, value=payment.get('account_number', ''))
            ws.cell(row=row_num, column=15, value=payment.get('sort_code', ''))
            ws.cell(row=row_num, column=16, value=payment.get('iban', ''))
            ws.cell(row=row_num, column=17, value=payment.get('swift_code', ''))
            ws.cell(row=row_num, column=18, value=payment.get('payment_reference', ''))
            ws.cell(row=row_num, column=19, value=payment.get('bank_address', ''))

        # Add total row
        total_row = len(invoices) + 2
        ws.cell(row=total_row, column=1, value='')
        ws.cell(row=total_row, column=6, value='TOTAL:')
        ws.cell(row=total_row, column=6).font = bold_font
        ws.cell(row=total_row, column=7, value=total_amount)
        ws.cell(row=total_row, column=7).font = bold_font

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(invoices) + 3):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'Invoice_Report_{today}.xlsx'

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        logger.info(f"Invoice report downloaded: {len(invoices)} invoices, total: {total_amount}")
        return response

    except Exception as e:
        logger.error(f"Error generating invoice report: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to generate report: {str(e)}'
        }), 500


# ========== API Routes ==========

@app.route('/api/invoices')
@handle_errors
@login_required
def api_get_invoices():
    """API: Get all invoices from Google Sheets"""
    try:
        invoices = get_invoices()
        return jsonify({
            'success': True,
            'data': invoices,
            'count': len(invoices)
        })
    except Exception as e:
        logger.error(f"API error getting invoices: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/<invoice_number>')
@handle_errors
@login_required
def api_get_invoice(invoice_number):
    """API: Get a specific invoice by number"""
    try:
        manager = get_sheets_manager()
        invoice = manager.get_invoice_by_number(invoice_number)

        if invoice:
            return jsonify({
                'success': True,
                'data': invoice
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Invoice {invoice_number} not found'
            }), 404
    except Exception as e:
        logger.error(f"API error getting invoice: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/stats')
@handle_errors
@login_required
def api_get_invoice_stats():
    """API: Get invoice statistics"""
    try:
        manager = get_sheets_manager()
        stats = manager.get_invoice_stats()
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        logger.error(f"API error getting stats: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/payment-details')
@handle_errors
@login_required
def api_get_payment_details():
    """API: Get all payment details from Google Sheets"""
    try:
        payments = get_payments()
        return jsonify({
            'success': True,
            'data': payments,
            'count': len(payments)
        })
    except Exception as e:
        logger.error(f"API error getting payment details: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/payment-details/<supplier_name>')
@handle_errors
@login_required
def api_get_payment_by_supplier(supplier_name):
    """API: Get payment details for a specific supplier"""
    try:
        manager = get_sheets_manager()
        payment = manager.get_payment_by_supplier(supplier_name)

        if payment:
            return jsonify({
                'success': True,
                'data': payment
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Payment details for {supplier_name} not found'
            }), 404
    except Exception as e:
        logger.error(f"API error getting payment details: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/payment-details/<supplier_name>', methods=['DELETE'])
@handle_errors
@login_required
def api_delete_payment_details(supplier_name):
    """API: Delete payment details for a specific supplier"""
    try:
        manager = get_sheets_manager()
        result = manager.delete_payment_by_supplier(supplier_name)

        if result.get('success'):
            logger.info(f"Payment details deleted for: {supplier_name}")
            return jsonify({
                'success': True,
                'message': f'Payment details for {supplier_name} deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('message', 'Failed to delete payment details')
            }), 404
    except Exception as e:
        logger.error(f"API error deleting payment details: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/suppliers')
@handle_errors
@login_required
def api_get_suppliers():
    """API: Get list of unique suppliers"""
    try:
        manager = get_sheets_manager()
        suppliers = manager.get_unique_suppliers()
        return jsonify({
            'success': True,
            'data': suppliers,
            'count': len(suppliers)
        })
    except Exception as e:
        logger.error(f"API error getting suppliers: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/<invoice_number>', methods=['DELETE'])
@handle_errors
@login_required
def api_delete_invoice(invoice_number):
    """API: Delete an invoice by number and its associated payment details"""
    try:
        manager = get_sheets_manager()

        # Get invoice first to retrieve file_id before deletion
        invoice = manager.get_invoice_by_number(invoice_number)
        file_id = invoice.get('file_id') if invoice else None

        # Delete the invoice
        result = manager.delete_invoice(invoice_number)

        if result.get('success'):
            logger.info(f"Invoice deleted: {invoice_number}")

            # Delete file from Google Drive if it exists
            if file_id:
                try:
                    drive_result = manager.delete_file_from_drive(file_id)
                    if drive_result.get('success'):
                        logger.info(f"File deleted from Google Drive: {file_id}")
                    else:
                        logger.warning(f"Failed to delete file from Drive: {drive_result.get('message')}")
                except Exception as e:
                    logger.warning(f"Failed to delete file from Google Drive: {e}")

            # Also delete associated payment details by invoice number
            try:
                payment_result = manager.delete_payment_by_invoice(invoice_number)
                if payment_result.get('success'):
                    logger.info(f"Payment details deleted for invoice: {invoice_number}")
            except Exception as e:
                logger.warning(f"Failed to delete payment details for invoice {invoice_number}: {e}")

            return jsonify({
                'success': True,
                'message': f'Invoice {invoice_number} and associated payment details deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('message', 'Failed to delete invoice')
            }), 404
    except Exception as e:
        logger.error(f"API error deleting invoice: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ========== Approve/Reject Routes ==========

@app.route('/api/invoices/approve-payments', methods=['POST'])
@handle_errors
@login_required
def api_approve_payments():
    """API: Approve selected invoices (change status to Approved)"""
    try:
        data = request.get_json() or {}
        invoice_numbers = data.get('invoice_numbers', [])

        if not invoice_numbers:
            return jsonify({
                'success': False,
                'error': 'No invoices selected'
            }), 400

        manager = get_sheets_manager()
        approved_count = 0

        for invoice_number in invoice_numbers:
            try:
                result = manager.update_invoice_status(invoice_number, 'Approved')
                if result.get('success'):
                    approved_count += 1
                    logger.info(f"Invoice approved: {invoice_number}")
                else:
                    logger.warning(f"Failed to approve invoice {invoice_number}: {result.get('message')}")
            except Exception as e:
                logger.warning(f"Failed to approve invoice {invoice_number}: {e}")

        return jsonify({
            'success': True,
            'approved_count': approved_count,
            'message': f'Successfully approved {approved_count} invoice(s)'
        })

    except Exception as e:
        logger.error(f"Error approving invoices: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/reject', methods=['POST'])
@handle_errors
@login_required
def api_reject_invoices():
    """API: Reject selected invoices (change status to Rejected)"""
    try:
        data = request.get_json() or {}
        invoice_numbers = data.get('invoice_numbers', [])

        if not invoice_numbers:
            return jsonify({
                'success': False,
                'error': 'No invoices selected'
            }), 400

        manager = get_sheets_manager()
        rejected_count = 0

        for invoice_number in invoice_numbers:
            try:
                result = manager.update_invoice_status(invoice_number, 'Rejected')
                if result.get('success'):
                    rejected_count += 1
                    logger.info(f"Invoice rejected: {invoice_number}")
                else:
                    logger.warning(f"Failed to reject invoice {invoice_number}: {result.get('message')}")
            except Exception as e:
                logger.warning(f"Failed to reject invoice {invoice_number}: {e}")

        return jsonify({
            'success': True,
            'rejected_count': rejected_count,
            'message': f'Successfully rejected {rejected_count} invoice(s)'
        })

    except Exception as e:
        logger.error(f"Error rejecting invoices: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/download-files', methods=['POST'])
@handle_errors
@login_required
def api_download_invoice_files():
    """API: Download original invoice files for selected invoices from Google Drive or local storage"""
    try:
        data = request.get_json() or {}
        invoice_numbers = data.get('invoice_numbers', [])

        if not invoice_numbers:
            return jsonify({
                'success': False,
                'error': 'No invoices selected'
            }), 400

        manager = get_sheets_manager()
        all_invoices = manager.get_all_invoices()

        # Find invoices with file_ids (Google Drive)
        found_files = []
        missing_file_ids = []
        for invoice in all_invoices:
            if invoice.get('invoice_number') in invoice_numbers:
                file_id = invoice.get('file_id')
                if file_id:
                    found_files.append({
                        'file_id': file_id,
                        'invoice_number': invoice.get('invoice_number'),
                        'source': 'drive'
                    })
                else:
                    missing_file_ids.append(invoice.get('invoice_number'))

        # Fall back to local files for invoices without file_id
        if missing_file_ids:
            uploads_folder = app.config['UPLOAD_FOLDER']
            for filename in os.listdir(uploads_folder):
                if filename.endswith('_data.json'):
                    json_path = os.path.join(uploads_folder, filename)
                    try:
                        with open(json_path, 'r') as f:
                            file_data = json.load(f)
                            if file_data.get('invoice_number') in missing_file_ids:
                                base_name = filename.replace('_data.json', '')
                                for ext in ['.pdf', '.png', '.jpg', '.jpeg']:
                                    invoice_file = base_name + ext
                                    invoice_path = os.path.join(uploads_folder, invoice_file)
                                    if os.path.exists(invoice_path):
                                        found_files.append({
                                            'path': invoice_path,
                                            'filename': invoice_file,
                                            'invoice_number': file_data.get('invoice_number'),
                                            'source': 'local'
                                        })
                                        break
                    except Exception as e:
                        logger.warning(f"Error reading {json_path}: {e}")

        if not found_files:
            return jsonify({
                'success': False,
                'error': 'No invoice files found for the selected invoices. Files may not have been uploaded to cloud storage.'
            }), 404

        # If only one file, download and send it directly
        if len(found_files) == 1:
            file_info = found_files[0]

            if file_info['source'] == 'drive':
                result = manager.download_file_from_drive(file_info['file_id'])
                if not result.get('success'):
                    return jsonify({
                        'success': False,
                        'error': result.get('message', 'Failed to download file')
                    }), 500
                file_buffer = io.BytesIO(result['content'])
                return send_file(
                    file_buffer,
                    mimetype=result['mime_type'],
                    as_attachment=True,
                    download_name=result['filename']
                )
            else:
                # Local file
                return send_file(
                    file_info['path'],
                    as_attachment=True,
                    download_name=file_info['filename']
                )

        # Multiple files - download all and create a zip
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for file_info in found_files:
                safe_inv_num = file_info['invoice_number'].replace('/', '-').replace('\\', '-')

                if file_info['source'] == 'drive':
                    result = manager.download_file_from_drive(file_info['file_id'])
                    if result.get('success'):
                        ext = os.path.splitext(result['filename'])[1] if result['filename'] else '.pdf'
                        archive_name = f"{safe_inv_num}{ext}"
                        zip_file.writestr(archive_name, result['content'])
                    else:
                        logger.warning(f"Failed to download file for invoice {file_info['invoice_number']}: {result.get('message')}")
                else:
                    # Local file
                    ext = os.path.splitext(file_info['filename'])[1]
                    archive_name = f"{safe_inv_num}{ext}"
                    zip_file.write(file_info['path'], archive_name)

        zip_buffer.seek(0)

        today = datetime.now().strftime('%Y-%m-%d')
        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=f'Invoices_{today}.zip'
        )

    except Exception as e:
        logger.error(f"Error downloading invoice files: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ========== Utility Routes ==========

@app.route('/api/test-connection')
@handle_errors
@login_required
def api_test_connection():
    """API: Test Google Sheets connection"""
    try:
        manager = get_sheets_manager()
        result = manager.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/initialize-sheets')
@handle_errors
@login_required
def api_initialize_sheets():
    """API: Initialize Google Sheets with headers"""
    try:
        manager = get_sheets_manager()
        result = manager.initialize_sheets()
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ========== Error Handlers ==========

@app.errorhandler(400)
def bad_request(e):
    """Handle bad request errors"""
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': 'Bad request',
            'message': str(e)
        }), 400
    flash('Bad request. Please check your input.', 'danger')
    return redirect(url_for('index'))


@app.errorhandler(404)
def not_found(e):
    """Handle not found errors"""
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': 'Not found'
        }), 404
    flash('Page not found.', 'warning')
    return redirect(url_for('index'))


@app.errorhandler(413)
def too_large(e):
    """Handle file too large error"""
    max_size_mb = app.config['MAX_CONTENT_LENGTH'] / (1024 * 1024)
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': f'File too large. Maximum size is {max_size_mb:.0f}MB'
        }), 413
    flash(f'File too large. Maximum size is {max_size_mb:.0f}MB.', 'danger')
    return redirect(url_for('upload'))


@app.errorhandler(500)
def server_error(e):
    """Handle server errors"""
    logger.error(f"Server error: {e}")
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500
    flash('An unexpected error occurred. Please try again.', 'danger')
    return redirect(url_for('index'))


# ========== Template Context ==========

@app.context_processor
def utility_processor():
    """Add utility functions to template context"""
    return {
        'now': datetime.now,
        'app_name': 'Invoice Tracker'
    }


# ========== Main Entry Point ==========

if __name__ == '__main__':
    # Print startup info
    print("=" * 50)
    print("Invoice Payment Tracker")
    print("=" * 50)
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    print(f"Max file size: {app.config['MAX_CONTENT_LENGTH'] / (1024*1024):.0f}MB")
    print(f"Allowed extensions: {', '.join(app.config['ALLOWED_EXTENSIONS'])}")
    print("=" * 50)

    # Test Google Sheets connection on startup
    try:
        manager = SheetsManager()
        conn_result = manager.test_connection()
        if conn_result['success']:
            print(f"Google Sheets: Connected to '{conn_result.get('spreadsheet_title')}'")
        else:
            print(f"Google Sheets: Connection failed - {conn_result.get('message')}")
    except Exception as e:
        print(f"Google Sheets: Not configured - {e}")

    print("=" * 50)
    print("Starting server at http://localhost:5000")
    print("=" * 50)

    # Run the app
    app.run(
        debug=os.getenv('FLASK_DEBUG', 'True').lower() == 'true',
        host='0.0.0.0',
        port=int(os.getenv('FLASK_PORT', 5000))
    )
