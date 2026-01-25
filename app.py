"""
Invoice Payment Tracker - Main Application
Handles invoice upload, OCR processing, and Google Sheets integration
"""

import os
import io
import json
import logging
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from functools import wraps
from flask import (
    Flask, render_template, request, jsonify,
    redirect, url_for, flash, session, make_response
)
from flask_cors import CORS
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

# Import our custom modules
from invoice_processor import InvoiceProcessor, process_invoice
from sheets_manager import (
    SheetsManager, SheetsManagerError, AuthenticationError,
    save_to_sheets, save_payment_details as save_payment_to_sheets,
    get_invoices, get_payments
)

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Initialize Flask app
app = Flask(__name__)
CORS(app)

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', os.urandom(24))
app.config['UPLOAD_FOLDER'] = os.path.abspath(os.getenv('UPLOAD_FOLDER', 'uploads'))
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_CONTENT_LENGTH', 16777216))  # 16MB
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'png', 'jpg', 'jpeg'}

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


# ========== Authentication ==========

def login_required(f):
    """Decorator to require login for routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('authenticated'):
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function


# ========== Helper Functions ==========

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


def get_file_extension(filename):
    """Get file extension"""
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else ''


def generate_unique_filename(original_filename):
    """Generate a unique filename with timestamp"""
    ext = get_file_extension(original_filename)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = secure_filename(original_filename.rsplit('.', 1)[0])
    return f"{base_name}_{timestamp}.{ext}"


def handle_errors(f):
    """Decorator for handling errors in routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except AuthenticationError as e:
            logger.error(f"Authentication error: {e}")
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'success': False,
                    'error': 'Google Sheets authentication failed. Please check your credentials.',
                    'error_type': 'authentication'
                }), 401
            flash('Google Sheets authentication failed. Please check your credentials.', 'danger')
            return redirect(url_for('index'))
        except SheetsManagerError as e:
            logger.error(f"Sheets manager error: {e}")
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'success': False,
                    'error': str(e),
                    'error_type': 'sheets_error'
                }), 500
            flash(f'Google Sheets error: {str(e)}', 'danger')
            return redirect(url_for('index'))
        except Exception as e:
            logger.error(f"Unexpected error: {e}", exc_info=True)
            if request.is_json or request.headers.get('Accept') == 'application/json':
                return jsonify({
                    'success': False,
                    'error': 'An unexpected error occurred',
                    'error_type': 'server_error'
                }), 500
            flash('An unexpected error occurred. Please try again.', 'danger')
            return redirect(url_for('index'))
    return decorated_function


def get_sheets_manager():
    """Get or create a SheetsManager instance"""
    try:
        return SheetsManager()
    except Exception as e:
        logger.error(f"Failed to initialize SheetsManager: {e}")
        raise


# ========== Authentication Routes ==========

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if session.get('authenticated'):
        return redirect(url_for('index'))

    if request.method == 'POST':
        password = request.form.get('password', '')
        app_password = os.environ.get('APP_PASSWORD', '')

        if not app_password:
            flash('APP_PASSWORD not configured. Please set it in your .env file.', 'error')
            return render_template('login.html')

        if password == app_password:
            session['authenticated'] = True
            session.permanent = True
            logger.info("User logged in successfully")
            next_url = request.args.get('next')
            return redirect(next_url or url_for('index'))
        else:
            flash('Invalid password. Please try again.', 'error')
            logger.warning("Failed login attempt")

    return render_template('login.html')


@app.route('/logout')
def logout():
    """Logout and clear session"""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ========== Main Routes ==========

@app.route('/')
@handle_errors
@login_required
def index():
    """Home page with dashboard"""
    try:
        manager = get_sheets_manager()
        stats = manager.get_invoice_stats()
        recent_invoices = manager.get_recent_invoices(limit=5)
    except Exception as e:
        logger.warning(f"Could not load dashboard data: {e}")
        stats = {
            'total_invoices': 0,
            'paid': 0,
            'pending': 0,
            'overdue': 0
        }
        recent_invoices = []

    return render_template('index.html', stats=stats, recent_invoices=recent_invoices)


@app.route('/dashboard')
@handle_errors
@login_required
def dashboard():
    """Dashboard page showing all invoices"""
    try:
        manager = get_sheets_manager()
        invoices = manager.get_all_invoices()
        stats = manager.get_invoice_stats()
    except Exception as e:
        logger.warning(f"Could not load dashboard data: {e}")
        invoices = []
        stats = {}

    return render_template('index.html', stats=stats, recent_invoices=invoices)


# ========== Upload Routes ==========

@app.route('/upload', methods=['GET', 'POST'])
@handle_errors
@login_required
def upload():
    """Handle invoice upload and processing"""
    if request.method == 'POST':
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes.best == 'application/json'

        # Check if file was uploaded (support both 'file' and 'invoice' field names)
        file = request.files.get('file') or request.files.get('invoice')

        if not file:
            logger.warning("Upload attempted with no file")
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': 'No file uploaded'
                }), 400
            flash('No file uploaded', 'danger')
            return redirect(url_for('upload'))

        if file.filename == '':
            logger.warning("Upload attempted with empty filename")
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': 'No file selected'
                }), 400
            flash('No file selected', 'danger')
            return redirect(url_for('upload'))

        if not allowed_file(file.filename):
            logger.warning(f"Upload attempted with invalid file type: {file.filename}")
            error_msg = f'Invalid file type. Allowed types: {", ".join(app.config["ALLOWED_EXTENSIONS"])}'
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 400
            flash(error_msg, 'danger')
            return redirect(url_for('upload'))

        # Generate unique filename and save
        filename = generate_unique_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        try:
            file.save(filepath)
            logger.info(f"File saved: {filepath}")
        except Exception as e:
            logger.error(f"Failed to save file: {e}")
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': 'Failed to save uploaded file'
                }), 500
            flash('Failed to save uploaded file', 'danger')
            return redirect(url_for('upload'))

        # Process invoice with Claude API
        try:
            logger.info(f"Processing invoice: {filepath}")
            extracted_data = process_invoice(filepath)

            # Check for processing errors
            if extracted_data.get('error'):
                logger.warning(f"Invoice processing returned error: {extracted_data.get('error')}")
                flash(f"Warning: {extracted_data.get('error')}", 'warning')

            logger.info(f"Invoice processed successfully: {extracted_data.get('invoice_number', 'N/A')}")

            # Save extracted data to JSON file (session cookie is too small)
            json_filename = filename.rsplit('.', 1)[0] + '_data.json'
            json_filepath = os.path.join(app.config['UPLOAD_FOLDER'], json_filename)
            with open(json_filepath, 'w') as f:
                json.dump(extracted_data, f)
            logger.info(f"Extracted data saved to: {json_filepath}")

            # Store only the filename in session (small enough for cookie)
            session['invoice_filename'] = filename

            # Always redirect to review page for verification
            if is_ajax:
                return jsonify({
                    'success': True,
                    'data': extracted_data,
                    'filename': filename,
                    'redirect': url_for('review', filename=filename)
                })

            # Direct redirect for form submission
            return redirect(url_for('review', filename=filename))

        except Exception as e:
            logger.error(f"Invoice processing failed: {e}", exc_info=True)
            error_msg = f'Failed to process invoice: {str(e)}'
            if is_ajax:
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 500
            flash(error_msg, 'danger')
            return redirect(url_for('upload'))

    # GET request - show upload form
    return render_template('upload.html')


# ========== Review Routes ==========

def convert_date_for_form(date_str):
    """Convert date from DD/MM/YYYY to YYYY-MM-DD format for HTML date inputs"""
    if not date_str:
        return ''
    # If already in YYYY-MM-DD format, return as is
    if len(date_str) == 10 and date_str[4] == '-' and date_str[7] == '-':
        return date_str
    # Try to parse DD/MM/YYYY format
    try:
        parts = date_str.split('/')
        if len(parts) == 3:
            day, month, year = parts
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
    except (ValueError, IndexError):
        pass
    return date_str


@app.route('/review')
@app.route('/review/<filename>')
@handle_errors
@login_required
def review(filename=None):
    """Review and edit extracted invoice data"""

    # Check if editing an existing invoice from Google Sheets
    invoice_number = request.args.get('invoice_number')
    if invoice_number:
        logger.info(f"Loading existing invoice for editing: {invoice_number}")
        try:
            manager = get_sheets_manager()
            invoice = manager.get_invoice_by_number(invoice_number)

            if not invoice:
                flash(f'Invoice {invoice_number} not found.', 'warning')
                return redirect(url_for('dashboard'))

            # Get associated payment details
            payment_details = None
            all_payments = manager.get_all_payment_details()
            for payment in all_payments:
                if payment.get('invoice_number') == invoice_number:
                    payment_details = payment
                    break

            # Format invoice data for the template (convert dates for form inputs)
            invoice_data = {
                'invoice_number': invoice.get('invoice_number', ''),
                'supplier_name': invoice.get('supplier_name', ''),
                'contact_email': invoice.get('contact_email', ''),
                'contact_phone': invoice.get('contact_phone', ''),
                'invoice_date': convert_date_for_form(invoice.get('invoice_date', '')),
                'due_date': convert_date_for_form(invoice.get('due_date', '')),
                'amount': invoice.get('amount', 0),
                'currency': invoice.get('currency', 'GBP'),
                'status': invoice.get('status', ''),
                'payment_date': convert_date_for_form(invoice.get('payment_date', '')),
                'notes': invoice.get('notes', ''),
                'payment_details': payment_details,
                '_row_id': invoice.get('id')  # Row number for updates
            }

            return render_template('review.html', invoice=invoice_data, is_editing=True)

        except Exception as e:
            logger.error(f"Error loading invoice for editing: {e}")
            flash(f'Error loading invoice: {str(e)}', 'danger')
            return redirect(url_for('dashboard'))

    # Try to get data from JSON file or query params (for new uploads)
    filename = filename or session.get('invoice_filename') or request.args.get('filename')
    invoice_data = {}

    if filename:
        # Try to load extracted data from JSON file
        json_filename = filename.rsplit('.', 1)[0] + '_data.json'
        json_filepath = os.path.join(app.config['UPLOAD_FOLDER'], json_filename)

        if os.path.exists(json_filepath):
            try:
                with open(json_filepath, 'r') as f:
                    invoice_data = json.load(f)
                logger.info(f"Loaded extracted data from: {json_filepath}")
            except Exception as e:
                logger.warning(f"Could not load JSON data: {e}")

        invoice_data['file_path'] = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    if not invoice_data or not invoice_data.get('invoice_number'):
        flash('No invoice data found. Please upload an invoice first.', 'warning')
        return redirect(url_for('upload'))

    return render_template('review.html', invoice=invoice_data, is_editing=False)


@app.route('/save-invoice', methods=['POST'])
@handle_errors
@login_required
def save_invoice():
    """Save invoice data to Google Sheets (create new or update existing)"""
    # Get form data
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form.to_dict()

    # Check if this is an edit of an existing invoice
    is_editing = data.get('is_editing') == 'true' or data.get('is_editing') == True
    row_id = data.get('_row_id')

    # Validate required fields
    required_fields = ['invoice_number', 'supplier_name', 'amount']
    missing_fields = [f for f in required_fields if not data.get(f)]

    if missing_fields:
        return jsonify({
            'success': False,
            'error': f'Missing required fields: {", ".join(missing_fields)}'
        }), 400

    # Clean and prepare data
    invoice_data = {
        'invoice_number': data.get('invoice_number', '').strip(),
        'supplier_name': data.get('supplier_name', '').strip(),
        'contact_email': data.get('contact_email', '').strip(),
        'contact_phone': data.get('contact_phone', '').strip(),
        'invoice_date': data.get('invoice_date', ''),
        'due_date': data.get('due_date', ''),
        'amount': float(data.get('amount', 0) or 0),
        'currency': data.get('currency', 'GBP').upper(),
        'status': data.get('status', 'Pending Review'),
        'payment_date': data.get('payment_date', ''),
        'notes': data.get('notes', '').strip()
    }

    # Extract payment details if present
    payment_details = None
    if data.get('payment_details') and isinstance(data.get('payment_details'), dict):
        payment_details = data.get('payment_details')
    else:
        # Check for individual payment fields in the form data
        payment_fields = ['beneficiary_account_name', 'account_number', 'iban',
                         'sort_code', 'swift_code', 'bank_name', 'bank_address', 'payment_reference']
        has_payment_data = any(data.get(f) for f in payment_fields)
        if has_payment_data:
            payment_details = {
                'beneficiary_account_name': data.get('beneficiary_account_name', '').strip(),
                'account_number': data.get('account_number', '').strip(),
                'iban': data.get('iban', '').strip().upper(),
                'sort_code': data.get('sort_code', '').strip(),
                'swift_code': data.get('swift_code', '').strip().upper(),
                'bank_name': data.get('bank_name', '').strip(),
                'bank_address': data.get('bank_address', '').strip(),
                'payment_reference': data.get('payment_reference', '').strip()
            }

    # Save to Google Sheets (update or create)
    try:
        manager = get_sheets_manager()

        if is_editing and row_id:
            # Update existing invoice
            row_number = int(row_id)
            result = manager.update_invoice(row_number, invoice_data)
            logger.info(f"Updated existing invoice: {invoice_data['invoice_number']} at row {row_number}")
        else:
            # Create new invoice
            result = save_to_sheets(invoice_data)

        if result.get('success'):
            # Handle payment details
            has_payment_info = payment_details and any(payment_details.values())
            payment_data = {
                'invoice_number': invoice_data['invoice_number'],
                'supplier_name': invoice_data['supplier_name'],
                'beneficiary_account_name': payment_details.get('beneficiary_account_name', '') if payment_details else '',
                'account_number': payment_details.get('account_number', '') if payment_details else '',
                'iban': payment_details.get('iban', '') if payment_details else '',
                'sort_code': payment_details.get('sort_code', '') if payment_details else '',
                'swift_code': payment_details.get('swift_code', '') if payment_details else '',
                'bank_name': payment_details.get('bank_name', '') if payment_details else '',
                'bank_address': payment_details.get('bank_address', '') if payment_details else '',
                'payment_reference': payment_details.get('payment_reference', '') if payment_details else '',
                'status': 'Auto-populated' if has_payment_info else 'Pending Details',
                'upload_date': '',
                'notes': f'Auto-populated from invoice {invoice_data["invoice_number"]}' if has_payment_info else f'Awaiting payment details - from invoice {invoice_data["invoice_number"]}'
            }

            try:
                if is_editing:
                    # Check if payment details exist for this invoice and update them
                    existing_payment = None
                    all_payments = manager.get_all_payment_details()
                    for p in all_payments:
                        if p.get('invoice_number') == invoice_data['invoice_number']:
                            existing_payment = p
                            break

                    if existing_payment:
                        manager.update_payment_details(existing_payment['id'], payment_data)
                        logger.info(f"Payment details updated for invoice: {invoice_data['invoice_number']}")
                    else:
                        save_payment_to_sheets(payment_data)
                        logger.info(f"Payment details created for invoice: {invoice_data['invoice_number']}")
                else:
                    save_payment_to_sheets(payment_data)
                    logger.info(f"Payment details saved for supplier: {invoice_data['supplier_name']}")
            except Exception as e:
                logger.warning(f"Failed to save payment details: {e}")

            # Clear session data (only for new uploads)
            if not is_editing:
                session.pop('extracted_invoice', None)
                session.pop('invoice_filename', None)

            success_msg = 'Invoice updated successfully!' if is_editing else 'Invoice saved successfully!'
            logger.info(f"Invoice {'updated' if is_editing else 'saved'} to sheets: {invoice_data['invoice_number']}")

            if request.is_json:
                return jsonify({
                    'success': True,
                    'message': success_msg,
                    'redirect': url_for('dashboard')
                })
            flash(success_msg, 'success')
            return redirect(url_for('dashboard'))
        else:
            error_msg = result.get('message', 'Failed to save invoice')
            logger.error(f"Failed to save invoice: {error_msg}")

            if request.is_json:
                return jsonify({
                    'success': False,
                    'error': error_msg
                }), 500
            flash(error_msg, 'danger')
            return redirect(url_for('review'))

    except Exception as e:
        logger.error(f"Error saving invoice: {e}", exc_info=True)
        if request.is_json:
            return jsonify({
                'success': False,
                'error': str(e)
            }), 500
        flash(f'Error saving invoice: {str(e)}', 'danger')
        return redirect(url_for('review'))


# ========== Payment Details Routes ==========

@app.route('/payment-details', methods=['GET', 'POST'])
@handle_errors
@login_required
def payment_details():
    """Manage payment details"""
    if request.method == 'POST':
        # Handle form submission
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()

        # Use new_supplier if provided, otherwise use selected supplier
        supplier_name = data.get('new_supplier', '').strip() or data.get('supplier_name', '').strip()

        if not supplier_name:
            if request.is_json:
                return jsonify({
                    'success': False,
                    'error': 'Supplier name is required'
                }), 400
            flash('Supplier name is required', 'danger')
            return redirect(url_for('payment_details'))

        # Prepare payment data
        payment_data = {
            'invoice_number': data.get('invoice_number', '').strip(),
            'supplier_name': supplier_name,
            'beneficiary_account_name': data.get('beneficiary_account_name', '').strip(),
            'account_number': data.get('account_number', '').strip(),
            'iban': data.get('iban', '').strip().upper(),
            'sort_code': data.get('sort_code', '').strip(),
            'swift_code': data.get('swift_code', '').strip().upper(),
            'bank_name': data.get('bank_name', '').strip(),
            'bank_address': data.get('bank_address', '').strip(),
            'payment_reference': data.get('payment_reference', '').strip(),
            'status': data.get('status', 'pending'),
            'upload_date': data.get('upload_date', ''),
            'notes': data.get('notes', '').strip()
        }

        # Save to Google Sheets
        try:
            result = save_payment_to_sheets(payment_data)

            if result.get('success'):
                logger.info(f"Payment details saved for: {supplier_name}")

                if request.is_json:
                    return jsonify({
                        'success': True,
                        'message': 'Payment details saved successfully'
                    })
                flash('Payment details saved successfully!', 'success')
                return redirect(url_for('payment_details'))
            else:
                error_msg = result.get('message', 'Failed to save payment details')
                logger.error(f"Failed to save payment details: {error_msg}")

                if request.is_json:
                    return jsonify({
                        'success': False,
                        'error': error_msg
                    }), 500
                flash(error_msg, 'danger')
                return redirect(url_for('payment_details'))

        except Exception as e:
            logger.error(f"Error saving payment details: {e}", exc_info=True)
            if request.is_json:
                return jsonify({
                    'success': False,
                    'error': str(e)
                }), 500
            flash(f'Error saving payment details: {str(e)}', 'danger')
            return redirect(url_for('payment_details'))

    # GET request - show payment details page
    try:
        manager = get_sheets_manager()
        payment_list = manager.get_all_payment_details()
        suppliers = manager.get_unique_suppliers()
        invoices = manager.get_all_invoices()
    except Exception as e:
        logger.warning(f"Could not load payment details: {e}")
        payment_list = []
        suppliers = []
        invoices = []

    return render_template(
        'payment_details.html',
        payment_details=payment_list,
        suppliers=suppliers,
        invoices=invoices
    )


@app.route('/save-payment-details', methods=['POST'])
@handle_errors
@login_required
def save_payment_details_route():
    """Alternative endpoint for saving payment details"""
    return payment_details()


@app.route('/download-payment-report')
@handle_errors
@login_required
def download_payment_report():
    """Download all payment details as an Excel file"""
    try:
        manager = get_sheets_manager()
        payments = manager.get_all_payment_details()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payment Details"

        # Define headers
        headers = [
            'Invoice Number',
            'Supplier Name',
            'Beneficiary Account Name',
            'Account Number',
            'IBAN',
            'Sort Code',
            'SWIFT/BIC Code',
            'Bank Name',
            'Bank Address',
            'Payment Reference',
            'Status',
            'Upload Date',
            'Notes'
        ]

        # Write headers with bold formatting
        bold_font = Font(bold=True)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = bold_font

        # Write data rows
        for row_num, payment in enumerate(payments, 2):
            ws.cell(row=row_num, column=1, value=payment.get('invoice_number', ''))
            ws.cell(row=row_num, column=2, value=payment.get('supplier_name', ''))
            ws.cell(row=row_num, column=3, value=payment.get('beneficiary_account_name', ''))
            ws.cell(row=row_num, column=4, value=payment.get('account_number', ''))
            ws.cell(row=row_num, column=5, value=payment.get('iban', ''))
            ws.cell(row=row_num, column=6, value=payment.get('sort_code', ''))
            ws.cell(row=row_num, column=7, value=payment.get('swift_code', ''))
            ws.cell(row=row_num, column=8, value=payment.get('bank_name', ''))
            ws.cell(row=row_num, column=9, value=payment.get('bank_address', ''))
            ws.cell(row=row_num, column=10, value=payment.get('payment_reference', ''))
            ws.cell(row=row_num, column=11, value=payment.get('status', ''))
            ws.cell(row=row_num, column=12, value=payment.get('upload_date', ''))
            ws.cell(row=row_num, column=13, value=payment.get('notes', ''))

        # Auto-size columns
        for col in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in range(1, len(payments) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'Payment_Report_{today}.xlsx'

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'

        logger.info(f"Payment report downloaded: {len(payments)} records")
        return response

    except Exception as e:
        logger.error(f"Error generating payment report: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to generate report: {str(e)}'
        }), 500


# ========== API Routes ==========

@app.route('/api/invoices')
@handle_errors
@login_required
def api_get_invoices():
    """API: Get all invoices from Google Sheets"""
    try:
        invoices = get_invoices()
        return jsonify({
            'success': True,
            'data': invoices,
            'count': len(invoices)
        })
    except Exception as e:
        logger.error(f"API error getting invoices: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/<invoice_number>')
@handle_errors
@login_required
def api_get_invoice(invoice_number):
    """API: Get a specific invoice by number"""
    try:
        manager = get_sheets_manager()
        invoice = manager.get_invoice_by_number(invoice_number)

        if invoice:
            return jsonify({
                'success': True,
                'data': invoice
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Invoice {invoice_number} not found'
            }), 404
    except Exception as e:
        logger.error(f"API error getting invoice: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/stats')
@handle_errors
@login_required
def api_get_invoice_stats():
    """API: Get invoice statistics"""
    try:
        manager = get_sheets_manager()
        stats = manager.get_invoice_stats()
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        logger.error(f"API error getting stats: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/payment-details')
@handle_errors
@login_required
def api_get_payment_details():
    """API: Get all payment details from Google Sheets"""
    try:
        payments = get_payments()
        return jsonify({
            'success': True,
            'data': payments,
            'count': len(payments)
        })
    except Exception as e:
        logger.error(f"API error getting payment details: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/payment-details/<supplier_name>')
@handle_errors
@login_required
def api_get_payment_by_supplier(supplier_name):
    """API: Get payment details for a specific supplier"""
    try:
        manager = get_sheets_manager()
        payment = manager.get_payment_by_supplier(supplier_name)

        if payment:
            return jsonify({
                'success': True,
                'data': payment
            })
        else:
            return jsonify({
                'success': False,
                'error': f'Payment details for {supplier_name} not found'
            }), 404
    except Exception as e:
        logger.error(f"API error getting payment details: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/payment-details/<supplier_name>', methods=['DELETE'])
@handle_errors
@login_required
def api_delete_payment_details(supplier_name):
    """API: Delete payment details for a specific supplier"""
    try:
        manager = get_sheets_manager()
        result = manager.delete_payment_by_supplier(supplier_name)

        if result.get('success'):
            logger.info(f"Payment details deleted for: {supplier_name}")
            return jsonify({
                'success': True,
                'message': f'Payment details for {supplier_name} deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('message', 'Failed to delete payment details')
            }), 404
    except Exception as e:
        logger.error(f"API error deleting payment details: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/suppliers')
@handle_errors
@login_required
def api_get_suppliers():
    """API: Get list of unique suppliers"""
    try:
        manager = get_sheets_manager()
        suppliers = manager.get_unique_suppliers()
        return jsonify({
            'success': True,
            'data': suppliers,
            'count': len(suppliers)
        })
    except Exception as e:
        logger.error(f"API error getting suppliers: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/invoices/<invoice_number>', methods=['DELETE'])
@handle_errors
@login_required
def api_delete_invoice(invoice_number):
    """API: Delete an invoice by number and its associated payment details"""
    try:
        manager = get_sheets_manager()

        # Delete the invoice
        result = manager.delete_invoice(invoice_number)

        if result.get('success'):
            logger.info(f"Invoice deleted: {invoice_number}")

            # Also delete associated payment details by invoice number
            try:
                payment_result = manager.delete_payment_by_invoice(invoice_number)
                if payment_result.get('success'):
                    logger.info(f"Payment details deleted for invoice: {invoice_number}")
            except Exception as e:
                logger.warning(f"Failed to delete payment details for invoice {invoice_number}: {e}")

            return jsonify({
                'success': True,
                'message': f'Invoice {invoice_number} and associated payment details deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('message', 'Failed to delete invoice')
            }), 404
    except Exception as e:
        logger.error(f"API error deleting invoice: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ========== Utility Routes ==========

@app.route('/api/test-connection')
@handle_errors
@login_required
def api_test_connection():
    """API: Test Google Sheets connection"""
    try:
        manager = get_sheets_manager()
        result = manager.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/initialize-sheets')
@handle_errors
@login_required
def api_initialize_sheets():
    """API: Initialize Google Sheets with headers"""
    try:
        manager = get_sheets_manager()
        result = manager.initialize_sheets()
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ========== Error Handlers ==========

@app.errorhandler(400)
def bad_request(e):
    """Handle bad request errors"""
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': 'Bad request',
            'message': str(e)
        }), 400
    flash('Bad request. Please check your input.', 'danger')
    return redirect(url_for('index'))


@app.errorhandler(404)
def not_found(e):
    """Handle not found errors"""
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': 'Not found'
        }), 404
    flash('Page not found.', 'warning')
    return redirect(url_for('index'))


@app.errorhandler(413)
def too_large(e):
    """Handle file too large error"""
    max_size_mb = app.config['MAX_CONTENT_LENGTH'] / (1024 * 1024)
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': f'File too large. Maximum size is {max_size_mb:.0f}MB'
        }), 413
    flash(f'File too large. Maximum size is {max_size_mb:.0f}MB.', 'danger')
    return redirect(url_for('upload'))


@app.errorhandler(500)
def server_error(e):
    """Handle server errors"""
    logger.error(f"Server error: {e}")
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500
    flash('An unexpected error occurred. Please try again.', 'danger')
    return redirect(url_for('index'))


# ========== Template Context ==========

@app.context_processor
def utility_processor():
    """Add utility functions to template context"""
    return {
        'now': datetime.now,
        'app_name': 'Invoice Tracker'
    }


# ========== Main Entry Point ==========

if __name__ == '__main__':
    # Print startup info
    print("=" * 50)
    print("Invoice Payment Tracker")
    print("=" * 50)
    print(f"Upload folder: {app.config['UPLOAD_FOLDER']}")
    print(f"Max file size: {app.config['MAX_CONTENT_LENGTH'] / (1024*1024):.0f}MB")
    print(f"Allowed extensions: {', '.join(app.config['ALLOWED_EXTENSIONS'])}")
    print("=" * 50)

    # Test Google Sheets connection on startup
    try:
        manager = SheetsManager()
        conn_result = manager.test_connection()
        if conn_result['success']:
            print(f"Google Sheets: Connected to '{conn_result.get('spreadsheet_title')}'")
        else:
            print(f"Google Sheets: Connection failed - {conn_result.get('message')}")
    except Exception as e:
        print(f"Google Sheets: Not configured - {e}")

    print("=" * 50)
    print("Starting server at http://localhost:5000")
    print("=" * 50)

    # Run the app
    app.run(
        debug=os.getenv('FLASK_DEBUG', 'True').lower() == 'true',
        host='0.0.0.0',
        port=int(os.getenv('FLASK_PORT', 5000))
    )
